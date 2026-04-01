import openpyxl
import sys

filepath = r'C:\Users\Admin\Dropbox\03-Clients_Laps_Media\#1 RUNES ET MAGIE\Produits .xlsx'

tarots = [
    {
        "nom": "Le Tarot Interdit",
        "sous_cat": "Tarot",
        "prix": "29.95$",
        "quantite": "1x",
        "description": "Plongez dans les profondeurs des contes interdits. Chaque carte est impregnee de l'esprit envoutant des recits obscurs. 78 cartes (22 arcanes majeurs + 56 mineurs). 38 personnages illustres. Format 4.5\"x6\". Par Nathaniel Dunhaven."
    },
    {
        "nom": "Le Tarot Akashique",
        "sous_cat": "Tarot",
        "prix": "34.95$",
        "quantite": "1x",
        "description": "Transporte dans la grande Salle des Annales pour trouver reponses, reveler talents, victoires inattendues, allies et prosperite sans precedent. 62 cartes. ISBN 9782898030451. Par Sharon A. Klingler et Sandra Anne Taylor."
    },
    {
        "nom": "Le Tarot des Anges",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Jeu bienveillant et sans danger, mots positifs, illustrations de Steve A. Roberts. Guide detaillant la marche a suivre a toutes les etapes. 78 cartes + guide. ISBN 9782898088681. Par Doreen Virtue et Radleigh Valentine. RUPTURE DE STOCK."
    },
    {
        "nom": "Le Tarot de l'Ere du Verseau",
        "sous_cat": "Tarot",
        "prix": "29.95$",
        "quantite": "1x",
        "description": "Celebration de l'esprit inventif de l'ere du Verseau. Chaque carte mele elements traditionnels et avant-gardistes pour eveiller l'intuition. 78 cartes. Format 4.5\"x6\". ISBN 9782898171185. Par Luna Ravenheart."
    },
    {
        "nom": "Tarot La Prophetie des Sorcieres",
        "sous_cat": "Tarot",
        "prix": "29.95$",
        "quantite": "1x",
        "description": "Reveillez la sorciere qui sommeille en vous. Fusionne sagesse des sorcieres et mysticisme moderne. 78 cartes + livret. Format 4.5\"x6\". ISBN 9782898171246. Par Isabella Moretti."
    },
    {
        "nom": "Tarot des Chamans de Lumiere",
        "sous_cat": "Tarot",
        "prix": "29.95$",
        "quantite": "1x",
        "description": "Voyage spirituel unique guide par sagesse ancestrale. Cle pour deverrouiller les mysteres de l'ame et du cosmos. 78 cartes + livret 344 pages. Format 4.5\"x6\". ISBN 9782898171178. Par Jasper Etherwind."
    },
    {
        "nom": "Tarot Mysteria",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Voyage qui eclaire passe, present et avenir. Chaque tirage est un rituel sacre revelant verites profondes. 78 cartes + livret 336 pages. Format 4.5\"x6\". ISBN 9782898171215. Par Isabella Moretti."
    },
    {
        "nom": "Le Tarot 444",
        "sous_cat": "Tarot",
        "prix": "29.95$",
        "quantite": "1x",
        "description": "Guide par le nombre angelique 444, symbole de stabilite et d'encouragement divin. 78 cartes + livret 192 pages. Format 4.5\"x6\". ISBN 9782898171260. Par Isabella Moretti."
    },
    {
        "nom": "Tarot de Marseille",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Explorez symbolique, forces, couleurs, saisons, signes astrologiques, huiles essentielles, chakras et pierres de lithotherapie. 78 cartes + livret. ISBN 9782898170461. Par Anne-Sophie Casper."
    },
    {
        "nom": "Tarot des Enfants de la Lune",
        "sous_cat": "Tarot",
        "prix": "29.95$",
        "quantite": "1x",
        "description": "Voyage mystique ou sagesse antique rencontre enchantement celeste. Arcanes majeurs pour grandes epopees, arcanes mineurs pour le quotidien. 78 cartes + livret. Format 4.5\"x6\". ISBN 9782898171147. Par Morgane Celeste."
    },
    {
        "nom": "Le Tarot de la Bienveillance",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Tarot classique teinte de psychologie positive. Messages sous forme d'affirmations positives. 4 elements : Air, Eau, Terre, Feu. 78 cartes. ISBN 9782897861384. Par Colette Baron-Reid. RUPTURE DE STOCK."
    },
    {
        "nom": "Tarot du Mirage Eternel",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Voyage a travers les dunes dorees de la conscience. Outil divinatoire melant tradition et innovation. 78 cartes + livret. Format 5\"x6.75\". ISBN 9782898171116. Par Jasper Etherwind."
    },
    {
        "nom": "Tarot du Sang de l'Ombre",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Monde ou lumiere rencontre tenebres. Esthetique gothique captivante, symboles mystiques. 78 cartes + livret 336 pages. Format 4.5\"x6\". ISBN 9782898171208. Par Nathaniel Dunhaven."
    },
    {
        "nom": "Tarot Dore - Guide Pratique",
        "sous_cat": "Tarot",
        "prix": "9.95$",
        "quantite": "1x",
        "description": "Code d'acces aux images enchanteresses du Tarot Dore. Enseignement intuitif des illustrations. Guide pratique. ISBN 9782898087813. Par Barbara Moore et Ciro Marchetti."
    },
    {
        "nom": "Le Tarot Simplifie",
        "sous_cat": "Tarot",
        "prix": "39.95$",
        "quantite": "1x",
        "description": "Coffret debutants : livre convivial, Tarot Dore, grille de tirage, conseils, raccourcis et exemples. Livre + Tarot + grille. ISBN 9782898086830. Par Josephine Ellershaw et Ciro Marchetti."
    },
    {
        "nom": "Tarot des Fees",
        "sous_cat": "Tarot",
        "prix": "39.95$",
        "quantite": "1x",
        "description": "Fees comme anges de la nature aidant a porter sa couronne invisible. Jeu de l'estime de soi. Illustrations de Howard David Johnson. 78 cartes + guide. ISBN 9782897676384. Par Doreen Virtue et Radleigh Valentine."
    },
    {
        "nom": "Le Tarot des Sorcieres",
        "sous_cat": "Tarot",
        "prix": "49.95$",
        "quantite": "1x",
        "description": "Base sur Rider-Waite-Smith, sorcellerie au devant de la scene. Illustrations de Mark Evans. Tirages : Triple Deesse, quatre elements, Roue de l'annee. 78 cartes + livre 312 pages. ISBN 9782898087066. Par Ellen Dugan."
    },
    {
        "nom": "Le Tarot du Pouvoir des Archanges",
        "sous_cat": "Tarot",
        "prix": "39.95$",
        "quantite": "1x",
        "description": "Messages precis et bienveillants des archanges. Concu pour gens de grande sensibilite ayant besoin d'encouragement. 78 cartes + livret. ISBN 9782897339104. Par Doreen Virtue et Radleigh Valentine."
    },
    {
        "nom": "Le Tarot Psychique - Cartes Oracles",
        "sous_cat": "Tarot",
        "prix": "34.95$",
        "quantite": "1x",
        "description": "Pont entre facultes psychiques et tarot. Developpement de l'intuition. Techniques : couleurs, symbolique, formes, numerologie, chakras. 65 cartes + guide. ISBN 9782896670925. Par John Holland."
    },
    {
        "nom": "Tarot Classique",
        "sous_cat": "Tarot",
        "prix": "39.95$",
        "quantite": "1x",
        "description": "Interpretation fidele du systeme Rider-Waite-Smith. Analyse detaillee, symbolisme decode, mots cles, associations astrologiques et elementales. 78 cartes + livre. ISBN 9782897527945. Par Barbara Moore et Eugene Smith."
    },
]

try:
    wb = openpyxl.load_workbook(filepath)
except Exception as e:
    print(f"ERREUR: Impossible d'ouvrir le fichier. Fermez Excel d'abord. ({e})")
    sys.exit(1)

ws = wb['Cartes divinatoires']

# Effacer les anciennes donnees (lignes 2+)
for row in range(ws.max_row, 1, -1):
    ws.delete_rows(row)

# Ecrire les nouveaux produits
for i, t in enumerate(tarots):
    row = i + 2
    ws.cell(row=row, column=1, value="Cartes divinatoires")
    ws.cell(row=row, column=2, value=t["sous_cat"])
    ws.cell(row=row, column=3, value=t["nom"])
    ws.cell(row=row, column=4, value=t["prix"])
    ws.cell(row=row, column=5, value="")  # lien images
    ws.cell(row=row, column=6, value="")  # En Stock
    ws.cell(row=row, column=7, value=t["quantite"])
    ws.cell(row=row, column=8, value=t["description"])

wb.save(filepath)
print(f"OK! {len(tarots)} tarots ecrits dans la feuille 'Cartes divinatoires'")
